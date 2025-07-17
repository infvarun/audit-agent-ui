from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
import os
import json
import openpyxl
from pydantic import BaseModel
import asyncio

from .database import get_db
from .storage import get_storage
from .langchain_service import get_question_analysis_service

router = APIRouter()

# Pydantic models for request/response validation
class ApplicationCreate(BaseModel):
    auditName: str
    name: str
    ciId: str
    startDate: str
    endDate: str
    settings: Dict[str, Any] = {}

class ApplicationResponse(BaseModel):
    id: int
    auditName: str
    name: str
    ciId: str
    startDate: str
    endDate: str
    settings: Dict[str, Any]
    createdAt: str

class DataRequestCreate(BaseModel):
    applicationId: int
    fileName: str
    filePath: str
    fileSize: int
    fileType: str
    categories: List[str] = []
    subcategories: List[str] = []
    columnMappings: Dict[str, Any] = {}
    questions: List[Dict[str, Any]] = []

class ToolConnectorCreate(BaseModel):
    applicationId: int
    ciId: str
    name: str
    type: str
    config: Dict[str, Any] = {}
    status: str = "pending"

class QuestionAnalysisCreate(BaseModel):
    applicationId: int
    analyses: List[Dict[str, Any]]

# Application routes
@router.get("/applications")
async def get_applications(db: Session = Depends(get_db), storage = Depends(get_storage)):
    """Get all applications"""
    try:
        applications = storage.get_all_applications(db)
        return applications
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to fetch applications")

@router.post("/applications")
async def create_application(
    application: ApplicationCreate,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Create new application"""
    try:
        app = storage.create_application(db, application.dict())
        return app
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/applications/{application_id}")
async def get_application(
    application_id: int,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Get specific application"""
    try:
        app = storage.get_application(db, application_id)
        if not app:
            raise HTTPException(status_code=404, detail="Application not found")
        return app
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to fetch application")

@router.put("/applications/{application_id}")
async def update_application(
    application_id: int,
    application: ApplicationCreate,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Update application"""
    try:
        app = storage.update_application(db, application_id, application.dict())
        if not app:
            raise HTTPException(status_code=404, detail="Application not found")
        return app
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Data request routes
@router.post("/data-requests")
async def create_data_request(
    data_request: DataRequestCreate,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Create data request"""
    try:
        request = storage.create_data_request(db, data_request.dict())
        return request
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/data-requests/application/{application_id}")
async def get_data_requests_by_application(
    application_id: int,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Get data requests for an application"""
    try:
        requests = storage.get_data_requests_by_application_id(db, application_id)
        return requests
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to fetch data requests")

# Excel file processing routes
@router.post("/excel/process")
async def process_excel_file(
    file: UploadFile = File(...),
    application_id: int = Form(...),
    file_type: str = Form(...),
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Process uploaded Excel file"""
    try:
        # Save uploaded file
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, file.filename)
        
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Process Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Extract data
        headers = [cell.value for cell in sheet[1]]
        data = []
        questions = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data)
            
            # Extract question if present
            if "Question" in row_data and row_data["Question"]:
                questions.append({
                    "id": str(len(questions)),
                    "question": row_data["Question"],
                    "category": row_data.get("Process", ""),
                    "subcategory": row_data.get("Sub-Process", "")
                })
        
        # Create data request
        data_request = {
            "applicationId": application_id,
            "fileName": file.filename,
            "filePath": file_path,
            "fileSize": len(content),
            "fileType": file_type,
            "categories": list(set(row.get("Process", "") for row in data if row.get("Process"))),
            "subcategories": list(set(row.get("Sub-Process", "") for row in data if row.get("Sub-Process"))),
            "columnMappings": {header: header for header in headers},
            "questions": questions
        }
        
        request = storage.create_data_request(db, data_request)
        
        return {
            "success": True,
            "dataRequest": request,
            "totalQuestions": len(questions),
            "categories": data_request["categories"],
            "subcategories": data_request["subcategories"]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process Excel file: {str(e)}")

@router.get("/excel/columns")
async def get_excel_columns(file_path: str):
    """Get column information from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        sample_data = []
        
        for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
            sample_data.append(dict(zip(headers, row)))
        
        return {
            "columns": headers,
            "sampleData": sample_data
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read Excel file: {str(e)}")

# Question analysis routes
@router.post("/questions/analyze")
async def analyze_questions(
    request: Dict[str, Any],
    db: Session = Depends(get_db),
    storage = Depends(get_storage),
    analysis_service = Depends(get_question_analysis_service)
):
    """Analyze questions using LangChain and OpenAI"""
    try:
        application_id = request.get("applicationId")
        if not application_id:
            raise HTTPException(status_code=400, detail="Application ID is required")
        
        # Get questions from data requests
        data_requests = storage.get_data_requests_by_application_id(db, application_id)
        if not data_requests:
            raise HTTPException(status_code=404, detail="No data requests found for this application")
        
        all_questions = []
        for req in data_requests:
            if req.questions:
                all_questions.extend(req.questions)
        
        if not all_questions:
            raise HTTPException(status_code=404, detail="No questions found to analyze")
        
        # Analyze questions using LangChain
        analyzed_questions = await analysis_service.analyze_questions_batch(all_questions)
        
        return {
            "success": True,
            "questions": analyzed_questions,
            "totalQuestions": len(analyzed_questions)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to analyze questions: {str(e)}")

@router.post("/questions/analyses/save")
async def save_question_analyses(
    request: QuestionAnalysisCreate,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Save question analyses to database"""
    try:
        # Delete existing analyses for this application
        storage.delete_question_analyses_by_application_id(db, request.applicationId)
        
        # Save new analyses
        saved_analyses = []
        for analysis in request.analyses:
            analysis_data = {
                "applicationId": request.applicationId,
                "questionId": analysis.get("id", ""),
                "originalQuestion": analysis.get("originalQuestion", ""),
                "category": analysis.get("category", ""),
                "subcategory": analysis.get("subcategory", ""),
                "aiPrompt": analysis.get("prompt", ""),
                "toolSuggestion": analysis.get("toolSuggestion", ""),
                "connectorReason": analysis.get("connectorReason", ""),
                "connectorToUse": analysis.get("connectorToUse", "")
            }
            
            saved_analysis = storage.create_question_analysis(db, analysis_data)
            saved_analyses.append(saved_analysis)
        
        return {
            "success": True,
            "analyses": saved_analyses
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save analyses: {str(e)}")

@router.get("/questions/analyses/{application_id}")
async def get_question_analyses(
    application_id: int,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Get saved question analyses"""
    try:
        analyses = storage.get_question_analyses_by_application_id(db, application_id)
        return {"analyses": analyses}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch analyses: {str(e)}")

# Tool connector routes
@router.post("/connectors")
async def create_tool_connector(
    connector: ToolConnectorCreate,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Create tool connector"""
    try:
        conn = storage.create_tool_connector(db, connector.dict())
        return conn
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/connectors/application/{application_id}")
async def get_connectors_by_application(
    application_id: int,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Get connectors for an application"""
    try:
        connectors = storage.get_tool_connectors_by_application_id(db, application_id)
        return connectors
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to fetch connectors")

@router.get("/connectors/ci/{ci_id}")
async def get_connectors_by_ci_id(
    ci_id: str,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Get connectors by CI ID"""
    try:
        connectors = storage.get_tool_connectors_by_ci_id(db, ci_id)
        return connectors
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to fetch connectors")

# Data collection routes
@router.post("/data-collection/start")
async def start_data_collection(
    request: Dict[str, Any],
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Start data collection session"""
    try:
        application_id = request.get("applicationId")
        if not application_id:
            raise HTTPException(status_code=400, detail="Application ID is required")
        
        # Create data collection session
        session_data = {
            "applicationId": application_id,
            "status": "running",
            "progress": 0,
            "logs": []
        }
        
        session = storage.create_data_collection_session(db, session_data)
        return session
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start data collection: {str(e)}")

@router.get("/data-collection/session/{application_id}")
async def get_data_collection_session(
    application_id: int,
    db: Session = Depends(get_db),
    storage = Depends(get_storage)
):
    """Get data collection session"""
    try:
        session = storage.get_data_collection_session_by_application_id(db, application_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        return session
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to fetch session")

# Health check
@router.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "python-backend"}